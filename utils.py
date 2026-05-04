import pandas as pd
import numpy as np
import re
import io
import base64
from datetime import date
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter


# ═══════════════════════════════════════════════════════════════════════════════
#  LEITURA DE ARQUIVOS
# ═══════════════════════════════════════════════════════════════════════════════

def ler_arquivo(uploaded_file) -> pd.DataFrame:
    nome = uploaded_file.name.lower()
    if nome.endswith((".xlsx", ".xls")):
        return pd.read_excel(uploaded_file, dtype=str, engine="openpyxl")
    for enc in ("utf-8", "latin-1", "cp1252"):
        try:
            uploaded_file.seek(0)
            return pd.read_csv(uploaded_file, dtype=str, sep=None, engine="python", encoding=enc)
        except (UnicodeDecodeError, UnicodeError):
            continue
    uploaded_file.seek(0)
    return pd.read_csv(uploaded_file, dtype=str, sep=None, engine="python", encoding="utf-8", errors="replace")


# ═══════════════════════════════════════════════════════════════════════════════
#  DETECÇÃO DE COLUNAS (suporta nomes truncados do Excel)
# ═══════════════════════════════════════════════════════════════════════════════

COLUNAS_CPF = ["cpf", "cpf_cnpj", "documento", "doc"]
COLUNAS_NOME = ["nome", "name", "cliente", "nome_cliente", "nome completo"]
COLUNAS_TELEFONE = ["telefone", "tel", "celular", "phone", "fone", "whatsapp"]
COLUNAS_VALOR_LIBERADO = ["valor liberado", "valor liber", "valor_liberado", "vlr_liberado", "valor", "valor_venda", "value", "receita"]
COLUNAS_BANCO = ["banco", "bank", "instituicao", "instituição"]
COLUNAS_PRODUTO = ["produto", "product", "tipo_produto", "tipo produto"]
COLUNAS_EQUIPE = ["equipe", "team", "time", "grupo"]
COLUNAS_VENDEDOR = ["vendedor", "seller", "consultor", "operador"]
COLUNAS_DATA_CADASTRO = ["data cadastro", "data cada", "data_cadastro", "dt_cadastro", "data"]
COLUNAS_DATA_NASCIMENTO = ["data de nascimento", "data de na", "data_nascimento", "dt_nascimento", "nascimento"]
COLUNAS_ID = ["id", "id_venda", "codigo", "código"]


def encontrar_coluna(df: pd.DataFrame, candidatas: list[str]) -> Optional[str]:
    colunas_norm = {c.lower().strip(): c for c in df.columns}
    for candidata in candidatas:
        cand_lower = candidata.lower()
        if cand_lower in colunas_norm:
            return colunas_norm[cand_lower]
    for candidata in candidatas:
        cand_lower = candidata.lower()
        for col_lower, col_original in colunas_norm.items():
            if col_lower.startswith(cand_lower) or cand_lower.startswith(col_lower):
                return col_original
    return None


# ═══════════════════════════════════════════════════════════════════════════════
#  NORMALIZAÇÃO
# ═══════════════════════════════════════════════════════════════════════════════

def normalizar_cpf(valor) -> Optional[str]:
    if pd.isna(valor) or valor is None:
        return None
    texto = str(valor).strip()
    try:
        f = float(texto)
        if f > 0:
            texto = str(int(f))
    except (ValueError, OverflowError):
        pass
    texto = re.sub(r"\D", "", texto)
    if len(texto) == 0:
        return None
    texto = texto.zfill(11)
    if len(texto) > 11:
        return None
    return texto


def normalizar_telefone(valor) -> Optional[str]:
    if pd.isna(valor) or valor is None:
        return None
    texto = str(valor).strip()
    try:
        f = float(texto)
        if f > 0:
            texto = str(int(f))
    except (ValueError, OverflowError):
        pass
    texto = re.sub(r"\D", "", texto)
    if len(texto) < 10:
        return None
    if len(texto) == 13 and texto.startswith("55"):
        texto = texto[2:]
    if len(texto) == 12 and texto.startswith("0"):
        texto = texto[1:]
    return texto


_MAPA_BANCOS = [
    (r"v8", "V8"),
    (r"c6", "C6"),
    (r"mercantil", "Mercantil"),
    (r"presen[cç]a", "Presença"),
    (r"facta", "Facta"),
]

def normalizar_banco(valor) -> str:
    if pd.isna(valor) or valor is None:
        return "Não informado"
    texto = str(valor).strip()
    if texto == "":
        return "Não informado"
    for pattern, nome in _MAPA_BANCOS:
        if re.search(pattern, texto, re.IGNORECASE):
            return nome
    return texto


_MAPA_PRODUTOS = [
    (r"fgts", "FGTS"),
    (r"cr[eé]dito\s+do\s+trabalhador", "CLT"),
    (r"\bclt\b", "CLT"),
]

def normalizar_produto(valor) -> str:
    if pd.isna(valor) or valor is None:
        return "Não informado"
    texto = str(valor).strip()
    if texto == "":
        return "Não informado"
    for pattern, nome in _MAPA_PRODUTOS:
        if re.search(pattern, texto, re.IGNORECASE):
            return nome
    return texto


def normalizar_valor(valor) -> float:
    if pd.isna(valor) or valor is None:
        return 0.0
    texto = str(valor).strip()
    texto = re.sub(r"[R$\s]", "", texto)
    if texto == "":
        return 0.0
    if re.match(r"^-?\d{1,3}(\.\d{3})*(,\d+)?$", texto):
        texto = texto.replace(".", "").replace(",", ".")
    elif "," in texto and "." in texto:
        if texto.rfind(",") > texto.rfind("."):
            texto = texto.replace(".", "").replace(",", ".")
        else:
            texto = texto.replace(",", "")
    elif "," in texto:
        texto = texto.replace(",", ".")
    try:
        return float(texto)
    except ValueError:
        return 0.0


def parse_data(valor) -> Optional[pd.Timestamp]:
    if pd.isna(valor) or valor is None:
        return pd.NaT
    texto = str(valor).strip()
    if texto == "":
        return pd.NaT
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return pd.Timestamp(pd.to_datetime(texto, format=fmt))
        except (ValueError, TypeError):
            continue
    try:
        return pd.Timestamp(pd.to_datetime(texto, dayfirst=True))
    except (ValueError, TypeError):
        return pd.NaT


def calcular_idade(data_nascimento, data_referencia=None) -> Optional[int]:
    if pd.isna(data_nascimento):
        return None
    if data_referencia is None:
        data_referencia = pd.Timestamp(date.today())
    try:
        idade = data_referencia.year - data_nascimento.year
        if (data_referencia.month, data_referencia.day) < (data_nascimento.month, data_nascimento.day):
            idade -= 1
        if 0 < idade < 120:
            return idade
        return None
    except (AttributeError, TypeError):
        return None


def faixa_etaria(idade) -> str:
    if pd.isna(idade) or idade is None:
        return "Não informado"
    idade = int(idade)
    if idade < 25:
        return "<25"
    elif idade < 35:
        return "25-34"
    elif idade < 45:
        return "35-44"
    elif idade < 55:
        return "45-54"
    elif idade < 65:
        return "55-64"
    else:
        return "65+"


# ═══════════════════════════════════════════════════════════════════════════════
#  VALIDAÇÃO
# ═══════════════════════════════════════════════════════════════════════════════

def validar_planilha_leads(df: pd.DataFrame) -> tuple[bool, str, list[str]]:
    warnings = []
    if df.empty:
        return False, "A planilha de leads está vazia.", warnings
    col_cpf = encontrar_coluna(df, COLUNAS_CPF)
    col_tel = encontrar_coluna(df, COLUNAS_TELEFONE)
    if col_cpf is None and col_tel is None:
        return False, (
            f"A planilha de leads não possui coluna de CPF nem de telefone. "
            f"Colunas encontradas: {', '.join(df.columns.tolist())}"
        ), warnings
    if col_cpf is None:
        warnings.append("Coluna de CPF não encontrada — cruzamento será apenas por telefone.")
    if col_tel is None:
        warnings.append("Coluna de telefone não encontrada — sem fallback por telefone.")
    return True, "", warnings


def validar_planilha_vendas(df: pd.DataFrame) -> tuple[bool, str, list[str]]:
    warnings = []
    if df.empty:
        return False, "A planilha de vendas está vazia.", warnings
    col_cpf = encontrar_coluna(df, COLUNAS_CPF)
    col_tel = encontrar_coluna(df, COLUNAS_TELEFONE)
    if col_cpf is None and col_tel is None:
        return False, (
            f"A planilha de vendas não possui coluna de CPF nem de telefone. "
            f"Colunas encontradas: {', '.join(df.columns.tolist())}"
        ), warnings
    col_valor = encontrar_coluna(df, COLUNAS_VALOR_LIBERADO)
    if col_valor is None:
        return False, (
            f"A planilha de vendas não possui coluna de valor. "
            f"Colunas encontradas: {', '.join(df.columns.tolist())}"
        ), warnings
    if encontrar_coluna(df, COLUNAS_BANCO) is None:
        warnings.append("Coluna 'Banco' não encontrada — seção por banco ficará vazia.")
    if encontrar_coluna(df, COLUNAS_PRODUTO) is None:
        warnings.append("Coluna 'Produto' não encontrada — seção por produto ficará vazia.")
    if encontrar_coluna(df, COLUNAS_EQUIPE) is None:
        warnings.append("Coluna 'Equipe' não encontrada.")
    if encontrar_coluna(df, COLUNAS_VENDEDOR) is None:
        warnings.append("Coluna 'Vendedor' não encontrada.")
    if encontrar_coluna(df, COLUNAS_DATA_CADASTRO) is None:
        warnings.append("Coluna 'Data Cadastro' não encontrada — análise temporal ficará vazia.")
    if encontrar_coluna(df, COLUNAS_DATA_NASCIMENTO) is None:
        warnings.append("Coluna 'Data Nascimento' não encontrada — perfil etário ficará vazio.")
    return True, "", warnings


# ═══════════════════════════════════════════════════════════════════════════════
#  PREPARAÇÃO DOS DADOS
# ═══════════════════════════════════════════════════════════════════════════════

def preparar_leads(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    col_cpf = encontrar_coluna(df, COLUNAS_CPF)
    col_tel = encontrar_coluna(df, COLUNAS_TELEFONE)
    col_nome = encontrar_coluna(df, COLUNAS_NOME)

    df["_cpf"] = df[col_cpf].apply(normalizar_cpf) if col_cpf else None
    df["_telefone"] = df[col_tel].apply(normalizar_telefone) if col_tel else None
    df["_nome"] = df[col_nome].astype(str).str.strip() if col_nome else ""
    return df


def preparar_vendas(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    df = df.copy()
    col_cpf = encontrar_coluna(df, COLUNAS_CPF)
    col_tel = encontrar_coluna(df, COLUNAS_TELEFONE)
    col_nome = encontrar_coluna(df, COLUNAS_NOME)
    col_valor = encontrar_coluna(df, COLUNAS_VALOR_LIBERADO)
    col_banco = encontrar_coluna(df, COLUNAS_BANCO)
    col_produto = encontrar_coluna(df, COLUNAS_PRODUTO)
    col_equipe = encontrar_coluna(df, COLUNAS_EQUIPE)
    col_vendedor = encontrar_coluna(df, COLUNAS_VENDEDOR)
    col_data_cad = encontrar_coluna(df, COLUNAS_DATA_CADASTRO)
    col_data_nasc = encontrar_coluna(df, COLUNAS_DATA_NASCIMENTO)

    df["_cpf"] = df[col_cpf].apply(normalizar_cpf) if col_cpf else None
    df["_telefone"] = df[col_tel].apply(normalizar_telefone) if col_tel else None
    df["_nome"] = df[col_nome].astype(str).str.strip() if col_nome else ""

    valores_ignorados = 0
    if col_valor:
        df["_valor"] = df[col_valor].apply(normalizar_valor)
        valores_ignorados = int((df["_valor"] == 0).sum() & (df[col_valor].notna()).sum())
    else:
        df["_valor"] = 0.0

    df["_banco"] = df[col_banco].apply(normalizar_banco) if col_banco else "Não informado"
    df["_produto"] = df[col_produto].apply(normalizar_produto) if col_produto else "Não informado"
    df["_equipe"] = df[col_equipe].astype(str).str.strip() if col_equipe else "Não informado"
    df["_vendedor"] = df[col_vendedor].astype(str).str.strip() if col_vendedor else "Não informado"
    df["_data_cadastro"] = df[col_data_cad].apply(parse_data) if col_data_cad else pd.NaT
    df["_data_nascimento"] = df[col_data_nasc].apply(parse_data) if col_data_nasc else pd.NaT
    df["_idade"] = df["_data_nascimento"].apply(calcular_idade)
    df["_faixa_etaria"] = df["_idade"].apply(faixa_etaria)

    return df, valores_ignorados


# ═══════════════════════════════════════════════════════════════════════════════
#  CRUZAMENTO
# ═══════════════════════════════════════════════════════════════════════════════

def cruzar_leads_vendas(
    leads_df: pd.DataFrame, vendas_df: pd.DataFrame
) -> tuple[pd.DataFrame, pd.DataFrame]:
    leads = preparar_leads(leads_df)
    vendas, _ = preparar_vendas(vendas_df)

    leads["_convertido"] = False
    leads["_match_tipo"] = ""

    cpfs_vendas = set(vendas.loc[vendas["_cpf"].notna(), "_cpf"].unique())
    if cpfs_vendas:
        mask = leads["_cpf"].notna() & leads["_cpf"].isin(cpfs_vendas)
        leads.loc[mask, "_convertido"] = True
        leads.loc[mask, "_match_tipo"] = "CPF"

    tels_vendas = set(vendas.loc[vendas["_telefone"].notna(), "_telefone"].unique())
    if tels_vendas:
        nao_conv = ~leads["_convertido"]
        mask = nao_conv & leads["_telefone"].notna() & leads["_telefone"].isin(tels_vendas)
        leads.loc[mask, "_convertido"] = True
        leads.loc[mask, "_match_tipo"] = "Telefone"

    return leads, vendas


# ═══════════════════════════════════════════════════════════════════════════════
#  GERAÇÃO DE EXCEL
# ═══════════════════════════════════════════════════════════════════════════════

_VERDE_FILL = PatternFill(start_color="3FB57E", end_color="3FB57E", fill_type="solid")
_ZEBRA_FILL = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
_HEADER_FONT = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
_DATA_FONT = Font(name="Calibri", size=11)
_LABEL_FONT = Font(name="Calibri", size=11, bold=True, color="2D3436")
_VALUE_FONT = Font(name="Calibri", size=11, color="3FB57E")
_TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="2D3436")
_SUBTITLE_FONT = Font(name="Calibri", size=10, color="95A5A6")
_THIN_BORDER = Border(
    left=Side(style="thin", color="E8ECEF"),
    right=Side(style="thin", color="E8ECEF"),
    top=Side(style="thin", color="E8ECEF"),
    bottom=Side(style="thin", color="E8ECEF"),
)
_BRAND_COLORS_HEX = ["3FB57E", "2A8B5F", "1E6B47", "82E0AA", "D5F5E3", "48C9B0", "1ABC9C"]


def _write_df_to_ws(ws, df, start_row=1, currency_cols=None, pct_cols=None, date_cols=None):
    currency_cols = set(currency_cols or [])
    pct_cols = set(pct_cols or [])
    date_cols = set(date_cols or [])

    for c, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=c, value=str(col_name))
        cell.font = _HEADER_FONT
        cell.fill = _VERDE_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _THIN_BORDER

    df = df.loc[:, ~df.columns.duplicated()]

    for r, (_, row_data) in enumerate(df.iterrows(), start_row + 1):
        for c, col_name in enumerate(df.columns, 1):
            val = row_data[col_name]

            if isinstance(val, pd.Series):
                val = val.iloc[0] if len(val) > 0 else None

            try:
                if val is None or pd.isna(val):
                    val = None
                elif isinstance(val, (np.integer,)):
                    val = int(val)
                elif isinstance(val, (np.floating,)):
                    val = float(val)
                elif isinstance(val, pd.Timestamp):
                    val = val.to_pydatetime()
            except (ValueError, TypeError):
                pass

            if col_name in pct_cols and isinstance(val, (int, float)) and val is not None:
                val = val / 100

            cell = ws.cell(row=r, column=c, value=val)
            cell.font = _DATA_FONT
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical="center")

            if (r - start_row) % 2 == 0:
                cell.fill = _ZEBRA_FILL

            if col_name in currency_cols:
                cell.number_format = '"R$ "#,##0.00'
            elif col_name in pct_cols:
                cell.number_format = '0.0%'
            elif col_name in date_cols:
                cell.number_format = 'DD/MM/YYYY'

    for c in range(1, len(df.columns) + 1):
        max_len = len(str(df.columns[c - 1]))
        for r in range(start_row + 1, min(start_row + len(df) + 1, start_row + 100)):
            val = ws.cell(row=r, column=c).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(c)].width = min(max(max_len + 3, 12), 40)

    ws.freeze_panes = f"A{start_row + 1}"
    if len(df) > 0:
        last_col = get_column_letter(len(df.columns))
        ws.auto_filter.ref = f"A{start_row}:{last_col}{start_row + len(df)}"


def gerar_relatorio_excel(
    leads_df: pd.DataFrame,
    vendas_df: pd.DataFrame,
    resumo_geral: dict,
    por_banco: pd.DataFrame,
    por_produto: pd.DataFrame,
    banco_produto: pd.DataFrame,
    por_equipe: pd.DataFrame,
    por_vendedor: pd.DataFrame,
    temporal: pd.DataFrame,
    leads_convertidos: pd.DataFrame,
) -> bytes:
    wb = Workbook()
    _COL_RENAME = {"Valor_Total": "Valor Total"}

    # ── 1. Resumo Executivo ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumo Executivo"
    ws.merge_cells("A1:B1")
    title_cell = ws["A1"]
    title_cell.value = "Relatório de Performance - Grupo Digital"
    title_cell.font = _TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A2:B2")
    sub_cell = ws["A2"]
    sub_cell.value = f"Gerado em {date.today().strftime('%d/%m/%Y')}"
    sub_cell.font = _SUBTITLE_FONT
    sub_cell.alignment = Alignment(horizontal="center")
    row_idx = 4
    for key, val in resumo_geral.items():
        cell_label = ws.cell(row=row_idx, column=1, value=key)
        cell_label.font = _LABEL_FONT
        cell_label.border = _THIN_BORDER
        cell_value = ws.cell(row=row_idx, column=2)
        cell_value.font = _VALUE_FONT
        cell_value.alignment = Alignment(horizontal="right")
        cell_value.border = _THIN_BORDER
        if "Valor" in key or "Ticket" in key:
            cell_value.value = val
            cell_value.number_format = '"R$ "#,##0.00'
        elif "Taxa" in key:
            cell_value.value = val / 100 if isinstance(val, (int, float)) else val
            cell_value.number_format = '0.0%'
        else:
            cell_value.value = val
        if (row_idx - 4) % 2 == 1:
            cell_label.fill = _ZEBRA_FILL
            cell_value.fill = _ZEBRA_FILL
        row_idx += 1
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 32

    # ── 2. Por Banco ─────────────────────────────────────────────────────
    ws_b = wb.create_sheet("Por Banco")
    if not por_banco.empty:
        pb = por_banco.rename(columns=_COL_RENAME)
        _write_df_to_ws(ws_b, pb, currency_cols=["Valor Total", "Ticket Médio"], pct_cols=["% do Total"])
        nrows = len(pb)
        chart = BarChart()
        chart.title = "Valor por Banco"
        chart.style = 10
        data_ref = Reference(ws_b, min_col=3, min_row=1, max_row=nrows + 1)
        cats_ref = Reference(ws_b, min_col=1, min_row=2, max_row=nrows + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.series[0].graphicalProperties.solidFill = "3FB57E"
        chart.legend = None
        chart.width = 18
        chart.height = 12
        ws_b.add_chart(chart, "G2")
        pie = PieChart()
        pie.title = "Distribuição de Vendas"
        pie.style = 10
        data_ref2 = Reference(ws_b, min_col=2, min_row=1, max_row=nrows + 1)
        cats_ref2 = Reference(ws_b, min_col=1, min_row=2, max_row=nrows + 1)
        pie.add_data(data_ref2, titles_from_data=True)
        pie.set_categories(cats_ref2)
        for i in range(nrows):
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = _BRAND_COLORS_HEX[i % len(_BRAND_COLORS_HEX)]
            pie.series[0].data_points.append(pt)
        pie.width = 14
        pie.height = 12
        ws_b.add_chart(pie, "G17")
    else:
        ws_b["A1"] = "Sem dados disponíveis."
        ws_b["A1"].font = _DATA_FONT

    # ── 3. Por Produto ───────────────────────────────────────────────────
    ws_p = wb.create_sheet("Por Produto")
    if not por_produto.empty:
        pp = por_produto.rename(columns=_COL_RENAME)
        _write_df_to_ws(ws_p, pp, currency_cols=["Valor Total", "Ticket Médio"])
        nrows = len(pp)
        chart = BarChart()
        chart.title = "Valor por Produto"
        chart.style = 10
        data_ref = Reference(ws_p, min_col=3, min_row=1, max_row=nrows + 1)
        cats_ref = Reference(ws_p, min_col=1, min_row=2, max_row=nrows + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.series[0].graphicalProperties.solidFill = "3FB57E"
        chart.legend = None
        chart.width = 18
        chart.height = 12
        ws_p.add_chart(chart, "F2")
    else:
        ws_p["A1"] = "Sem dados disponíveis."
        ws_p["A1"].font = _DATA_FONT

    # ── 4. Banco x Produto ───────────────────────────────────────────────
    ws_bp = wb.create_sheet("Banco x Produto")
    if not banco_produto.empty:
        _write_df_to_ws(ws_bp, banco_produto, currency_cols=["Valor"])
    else:
        ws_bp["A1"] = "Sem dados disponíveis."
        ws_bp["A1"].font = _DATA_FONT

    # ── 5. Por Equipe ────────────────────────────────────────────────────
    ws_e = wb.create_sheet("Por Equipe")
    if not por_equipe.empty:
        pe = por_equipe.rename(columns=_COL_RENAME)
        _write_df_to_ws(ws_e, pe, currency_cols=["Valor Total", "Ticket Médio"])
    else:
        ws_e["A1"] = "Sem dados disponíveis."
        ws_e["A1"].font = _DATA_FONT

    # ── 6. Por Vendedor ──────────────────────────────────────────────────
    ws_v = wb.create_sheet("Por Vendedor")
    if not por_vendedor.empty:
        pv = por_vendedor.rename(columns=_COL_RENAME)
        _write_df_to_ws(ws_v, pv, currency_cols=["Valor Total", "Ticket Médio"])
        nrows = min(len(pv), 15)
        chart = BarChart()
        chart.title = "Top 15 Vendedores por Valor"
        chart.type = "bar"
        chart.style = 10
        data_ref = Reference(ws_v, min_col=3, min_row=1, max_row=nrows + 1)
        cats_ref = Reference(ws_v, min_col=1, min_row=2, max_row=nrows + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.series[0].graphicalProperties.solidFill = "3FB57E"
        chart.legend = None
        chart.width = 20
        chart.height = 14
        ws_v.add_chart(chart, "F2")
    else:
        ws_v["A1"] = "Sem dados disponíveis."
        ws_v["A1"].font = _DATA_FONT

    # ── 7. Análise Temporal ──────────────────────────────────────────────
    ws_t = wb.create_sheet("Análise Temporal")
    if not temporal.empty:
        _write_df_to_ws(ws_t, temporal, date_cols=["Data"], currency_cols=["Valor"])
        nrows = len(temporal)
        line = LineChart()
        line.title = "Vendas ao Longo do Tempo"
        line.style = 10
        line.y_axis.title = "Qtd Vendas"
        data_ref = Reference(ws_t, min_col=2, min_row=1, max_row=nrows + 1)
        cats_ref = Reference(ws_t, min_col=1, min_row=2, max_row=nrows + 1)
        line.add_data(data_ref, titles_from_data=True)
        line.set_categories(cats_ref)
        line.series[0].graphicalProperties.line.solidFill = "3FB57E"
        line.legend = None
        line.width = 20
        line.height = 12
        ws_t.add_chart(line, "E2")
        bar = BarChart()
        bar.title = "Valor ao Longo do Tempo"
        bar.style = 10
        bar.y_axis.title = "Valor (R$)"
        data_ref2 = Reference(ws_t, min_col=3, min_row=1, max_row=nrows + 1)
        cats_ref2 = Reference(ws_t, min_col=1, min_row=2, max_row=nrows + 1)
        bar.add_data(data_ref2, titles_from_data=True)
        bar.set_categories(cats_ref2)
        bar.series[0].graphicalProperties.solidFill = "2A8B5F"
        bar.legend = None
        bar.width = 20
        bar.height = 12
        ws_t.add_chart(bar, "E17")
    else:
        ws_t["A1"] = "Sem dados temporais disponíveis."
        ws_t["A1"].font = _DATA_FONT

    # ── 8. Perfil Etário ─────────────────────────────────────────────────
    ws_f = wb.create_sheet("Perfil Etário")
    vendas_com_idade = vendas_df[vendas_df["_idade"].notna()].copy() if "_idade" in vendas_df.columns else pd.DataFrame()
    if not vendas_com_idade.empty:
        faixa_order = ["<25", "25-34", "35-44", "45-54", "55-64", "65+"]
        faixas_df = vendas_com_idade.groupby("_faixa_etaria").agg(
            Quantidade=("_valor", "count"),
            Valor_Total=("_valor", "sum"),
        ).reindex(faixa_order).fillna(0).reset_index()
        faixas_df.columns = ["Faixa Etária", "Quantidade", "Valor Total"]
        faixas_df["Quantidade"] = faixas_df["Quantidade"].astype(int)
        _write_df_to_ws(ws_f, faixas_df, currency_cols=["Valor Total"])
        nrows = len(faixas_df)
        chart = BarChart()
        chart.title = "Distribuição por Faixa Etária"
        chart.style = 10
        data_ref = Reference(ws_f, min_col=2, min_row=1, max_row=nrows + 1)
        cats_ref = Reference(ws_f, min_col=1, min_row=2, max_row=nrows + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.series[0].graphicalProperties.solidFill = "3FB57E"
        chart.legend = None
        chart.width = 16
        chart.height = 12
        ws_f.add_chart(chart, "E2")
    else:
        ws_f["A1"] = "Sem dados de faixa etária disponíveis."
        ws_f["A1"].font = _DATA_FONT

    # ── 9. Leads Convertidos ─────────────────────────────────────────────
    ws_lc = wb.create_sheet("Leads Convertidos")
    if not leads_convertidos.empty:
        _write_df_to_ws(ws_lc, leads_convertidos)
    else:
        ws_lc["A1"] = "Nenhum lead convertido encontrado."
        ws_lc["A1"].font = _DATA_FONT

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
#  TABELA HTML CUSTOMIZADA (THEME-AWARE)
# ═══════════════════════════════════════════════════════════════════════════════

def render_tabela_html(
    df: pd.DataFrame,
    tema: str = "light",
    destacar_top3: bool = True,
    colunas_moeda: Optional[list] = None,
    colunas_percentual: Optional[list] = None,
) -> str:
    import hashlib
    uid = hashlib.md5(str(id(df)).encode() + str(len(df)).encode()).hexdigest()[:8]
    cls = f"tabela-customizada-{uid}"

    if tema == "dark":
        cores = {
            "bg_header": "#3FB57E",
            "text_header": "#FFFFFF",
            "bg_row_even": "#252930",
            "bg_row_odd": "#1A1D23",
            "text": "#E8E8E8",
            "border": "#3A3F47",
            "hover": "#2D3640",
        }
    else:
        cores = {
            "bg_header": "#3FB57E",
            "text_header": "#FFFFFF",
            "bg_row_even": "#F8F9FA",
            "bg_row_odd": "#FFFFFF",
            "text": "#2D3436",
            "border": "#E0E0E0",
            "hover": "#F0F7F3",
        }

    colunas_moeda = set(colunas_moeda or [])
    colunas_percentual = set(colunas_percentual or [])
    medalhas = {0: "🥇", 1: "🥈", 2: "🥉"}

    def fmt_val(col, val):
        if pd.isna(val):
            return ""
        if col in colunas_moeda:
            try:
                v = float(val)
                return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            except (ValueError, TypeError):
                return str(val)
        if col in colunas_percentual:
            try:
                return f"{float(val):.1f}%"
            except (ValueError, TypeError):
                return str(val)
        if isinstance(val, float):
            if val == int(val):
                return f"{int(val):,}".replace(",", ".")
            return f"{val:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        return str(val)

    def align_for(col):
        if col in colunas_moeda:
            return "right"
        if col in colunas_percentual:
            return "right"
        return "left"

    style_tag = (
        f"<style>"
        f".{cls} tr:hover td {{background-color: {cores['hover']} !important; transition: background-color 0.2s;}}"
        f"</style>"
    )

    header_cells = ""
    for col in df.columns:
        a = align_for(col)
        header_cells += (
            f'<th style="padding:14px 16px;text-align:{a};font-size:15px;'
            f'font-family:Inter,sans-serif;font-weight:600;'
            f'border-bottom:2px solid {cores["border"]};">{col}</th>'
        )

    body_rows = ""
    for i, (_, row) in enumerate(df.iterrows()):
        bg = cores["bg_row_even"] if i % 2 == 0 else cores["bg_row_odd"]
        cells = ""
        for j, col in enumerate(df.columns):
            a = align_for(col)
            val_str = fmt_val(col, row[col])
            if destacar_top3 and j == 0 and i in medalhas:
                val_str = f"{medalhas[i]} {val_str}"
            cells += (
                f'<td style="padding:12px 16px;text-align:{a};font-size:15px;'
                f'font-family:Inter,sans-serif;color:{cores["text"]};'
                f'border-bottom:1px solid {cores["border"]};background:{bg};">{val_str}</td>'
            )
        body_rows += f"<tr>{cells}</tr>"

    return (
        f'{style_tag}'
        f'<div style="border-radius:8px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,0.08);margin:1rem 0;">'
        f'<table class="{cls}" style="width:100%;border-collapse:collapse;">'
        f'<thead><tr style="background:{cores["bg_header"]};color:{cores["text_header"]};">'
        f'{header_cells}</tr></thead>'
        f'<tbody>{body_rows}</tbody>'
        f'</table></div>'
    )


# ═══════════════════════════════════════════════════════════════════════════════
#  EXPORTAÇÃO PDF
# ═══════════════════════════════════════════════════════════════════════════════

def _fmt_brl_pdf(valor) -> str:
    try:
        v = float(valor)
        return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except (ValueError, TypeError):
        return "R$ 0,00"


def _df_to_html_table(df: pd.DataFrame, max_rows: int = 0, tema: str = "light") -> str:
    is_dark = tema == "dark"
    header_bg = "#2A8B5F" if not is_dark else "#3FB57E"
    even_bg = "#F8F9FA" if not is_dark else "#2A2D35"
    odd_bg = "#FFFFFF" if not is_dark else "#252930"
    text_color = "#2D3436" if not is_dark else "#E8E8E8"
    border_color = "#E8ECEF" if not is_dark else "#3A3F47"

    rows = df.head(max_rows) if max_rows > 0 else df
    html = f'<table style="width:100%;border-collapse:collapse;font-size:10px;margin:10px 0;">'
    html += "<thead><tr>"
    for col in rows.columns:
        html += f'<th style="background:{header_bg};color:#FFF;padding:8px 10px;text-align:left;border:1px solid {border_color};">{col}</th>'
    html += "</tr></thead><tbody>"
    for i, (_, row) in enumerate(rows.iterrows()):
        bg = even_bg if i % 2 == 0 else odd_bg
        html += f'<tr style="background:{bg};">'
        for col in rows.columns:
            val = row[col]
            if isinstance(val, float):
                if "%" in col or "Taxa" in col or "Conversão" in col:
                    display = f"{val:.1f}%"
                elif "Valor" in col or "Ticket" in col or "Médio" in col:
                    display = _fmt_brl_pdf(val)
                else:
                    display = f"{val:,.0f}".replace(",", ".")
            else:
                display = str(val) if pd.notna(val) else ""
            html += f'<td style="padding:6px 10px;color:{text_color};border:1px solid {border_color};">{display}</td>'
        html += "</tr>"
    html += "</tbody></table>"
    return html


def _embed_chart(chart_images: dict, key: str) -> str:
    if key not in chart_images or not chart_images[key]:
        return ""
    b64 = base64.b64encode(chart_images[key]).decode("utf-8")
    return f'<div style="text-align:center;margin:15px 0;"><img src="data:image/png;base64,{b64}" style="max-width:100%;height:auto;" /></div>'


def gerar_relatorio_pdf(
    resumo_geral: dict,
    por_banco: pd.DataFrame,
    por_produto: pd.DataFrame,
    por_equipe: pd.DataFrame,
    por_vendedor: pd.DataFrame,
    periodo_txt: str,
    tema: str = "light",
    chart_images: Optional[dict] = None,
) -> bytes:
    try:
        from xhtml2pdf import pisa
    except ImportError:
        raise ImportError("xhtml2pdf não está instalado. Execute: pip install xhtml2pdf")

    if chart_images is None:
        chart_images = {}

    is_dark = tema == "dark"
    verde = "#4ECB8E" if is_dark else "#3FB57E"
    verde_escuro = "#3FB57E" if is_dark else "#2A8B5F"
    texto = "#E8E8E8" if is_dark else "#2D3436"
    texto_sec = "#A0A0A0" if is_dark else "#95A5A6"
    fundo = "#1A1D23" if is_dark else "#FFFFFF"
    card_bg = "#252930" if is_dark else "#F8F9FA"
    borda = "#3A3F47" if is_dark else "#E8ECEF"

    data_rel = resumo_geral.get("Data do Relatório", "")

    total_leads = resumo_geral.get("Total de Leads", 0)
    convertidos = resumo_geral.get("Leads Convertidos", 0)
    total_vendas = resumo_geral.get("Total de Vendas", 0)
    taxa = resumo_geral.get("Taxa de Conversão (%)", 0)
    valor_total = resumo_geral.get("Valor Total Liberado", 0)
    ticket = resumo_geral.get("Ticket Médio", 0)
    bancos_ativos = resumo_geral.get("Bancos Ativos", 0)
    vendedores_ativos = resumo_geral.get("Vendedores Ativos", 0)

    def _metric_cell(label: str, value) -> str:
        return (
            f'<td style="width:25%;padding:6px;">'
            f'<div style="background:{card_bg};border:1px solid {borda};'
            f'border-left:4px solid {verde};padding:10px 14px;">'
            f'<div style="color:{texto_sec};font-size:8px;text-transform:uppercase;">{label}</div>'
            f'<div style="color:{verde};font-size:16px;font-weight:bold;">{value}</div>'
            f'</div></td>'
        )

    metrics_html = (
        '<table style="width:100%;border-collapse:collapse;margin:12px 0;"><tr>'
        + _metric_cell("Total de Leads", f"{total_leads:,}")
        + _metric_cell("Leads Convertidos", f"{convertidos:,}")
        + _metric_cell("Taxa de Conversão", f"{taxa:.1f}%")
        + _metric_cell("Total de Vendas", f"{total_vendas:,}")
        + '</tr><tr>'
        + _metric_cell("Valor Liberado", _fmt_brl_pdf(valor_total))
        + _metric_cell("Ticket Médio", _fmt_brl_pdf(ticket))
        + _metric_cell("Bancos Ativos", str(bancos_ativos))
        + _metric_cell("Vendedores Ativos", str(vendedores_ativos))
        + '</tr></table>'
    )

    banco_table = _df_to_html_table(por_banco, tema=tema)
    pie_chart = _embed_chart(chart_images, "pie_banco")
    produto_table = _df_to_html_table(por_produto, tema=tema)
    equipe_table = _df_to_html_table(por_equipe, tema=tema)
    vendedor_table = _df_to_html_table(por_vendedor, max_rows=15, tema=tema)
    temporal_chart = _embed_chart(chart_images, "temporal")
    heatmap_chart = _embed_chart(chart_images, "heatmap")
    faixa_chart = _embed_chart(chart_images, "faixa_etaria")

    css = f"""
    @page {{
        size: A4 portrait;
        margin: 20mm 18mm 25mm 18mm;
    }}
    body {{
        font-family: Helvetica, Arial, sans-serif;
        color: {texto};
        background: {fundo};
        font-size: 10px;
        line-height: 1.4;
    }}
    .header {{
        text-align: center;
        color: {texto_sec};
        font-size: 8px;
        border-bottom: 1px solid {borda};
        padding-bottom: 6px;
        margin-bottom: 10px;
    }}
    .cover {{
        text-align: center;
        padding: 60px 0 30px 0;
        border-bottom: 3px solid {verde};
        margin-bottom: 20px;
    }}
    .cover h1 {{
        color: {verde};
        font-size: 26px;
        margin-bottom: 4px;
    }}
    .cover h2 {{
        color: {texto};
        font-size: 14px;
        font-weight: normal;
        margin-bottom: 8px;
    }}
    .cover .subtitle {{
        color: {texto_sec};
        font-size: 11px;
    }}
    .section-title {{
        color: {verde};
        font-size: 14px;
        font-weight: bold;
        border-bottom: 2px solid {verde};
        padding-bottom: 4px;
        margin: 20px 0 10px 0;
    }}
    .footer {{
        text-align: center;
        color: {texto_sec};
        font-size: 8px;
        border-top: 1px solid {borda};
        padding-top: 6px;
        margin-top: 15px;
    }}
    """

    sections_page2 = f"""
    <div class="section-title">Performance por Banco</div>
    {banco_table}
    {pie_chart}

    <div class="section-title">Performance por Produto</div>
    {produto_table}
    """

    sections_page3 = f"""
    <div class="section-title">Performance por Equipe</div>
    {equipe_table}

    <div class="section-title">Top 15 Vendedores</div>
    {vendedor_table}
    """

    sections_charts = ""
    if temporal_chart:
        sections_charts += f'<div class="section-title">Evolução Temporal</div>{temporal_chart}'
    if heatmap_chart:
        sections_charts += f'<div class="section-title">Banco x Produto</div>{heatmap_chart}'
    if faixa_chart:
        sections_charts += f'<div class="section-title">Faixa Etária</div>{faixa_chart}'

    html_content = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8" />
    <style>{css}</style>
</head>
<body>
    <div class="header">Grupo Digital - Soluções Financeiras | {data_rel}</div>

    <div class="cover">
        <h1>Grupo Digital</h1>
        <h2>Relatório de Performance de Leads</h2>
        <div class="subtitle">{periodo_txt} &middot; Gerado em {data_rel}</div>
    </div>

    <div class="section-title">Indicadores Gerais</div>
    {metrics_html}

    <pdf:nextpage />

    {sections_page2}

    <pdf:nextpage />

    {sections_page3}

    {"<pdf:nextpage />" + sections_charts if sections_charts else ""}

    <div class="footer">
        Grupo Digital - Soluções Financeiras | Relatório gerado automaticamente
    </div>
</body>
</html>"""

    buffer = io.BytesIO()
    status = pisa.CreatePDF(html_content, dest=buffer, encoding="utf-8")
    if status.err:
        raise RuntimeError("Erro ao gerar PDF. Verifique os dados do relatório.")
    return buffer.getvalue()
